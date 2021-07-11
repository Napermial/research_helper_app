import openpyxl
from io import BytesIO
from .models import Factor, Level, Item, ItemLevel, Experiment


class Line:
    """class to keep track of the lines of the read excel document"""

    def __init__(self, stimulus, pre_context, post_context, lexicalization, levels):
        self.stimulus = stimulus
        self.pre_context = pre_context
        self.post_context = post_context
        self.lexicalization = lexicalization
        self.levels = levels

    def __str__(self):
        return self.stimulus

    def __repr__(self):
        return self.stimulus


def read_file(input_file: object) -> (Line, list):
    """
    reads the file and composes the structure of the items
    :param input_file: the uploaded file
    :returns the lines as Line(s), the structure of the factors in the file
    """
    factors_in_workbook = {}
    workbook = openpyxl.load_workbook(filename=BytesIO(input_file.read()), data_only=True).active
    position = 4
    for header in workbook.iter_rows(min_row=1, max_row=1, values_only=True, min_col=5):
        for item in header:
            factors_in_workbook[item] = position
            position += 1
    lines = []
    for row in workbook.iter_rows(min_row=2, values_only=True):
        lines.append(Line(row[0], row[1], row[2], row[3], row[4:]))
    return lines, factors_in_workbook


def create_item(line: Line) -> Item:
    """
    creates an Item and saves it to the database
    :param line :  a Line object containing the data of the line
    :returns the Line to add the connected Level
    """
    item = Item()
    item.item_text = line.stimulus
    item.pre_item_context = line.pre_context
    item.post_item_context = line.post_context
    item.lexicalization = line.lexicalization
    item.save()
    return item


def put_into_db(file: dict, levels: list, experiment: Experiment):
    """
    inserts the lines to the database
    :param file: the read file
    :param levels: list of the created levels with metadata
    :param experiment: the Experiment object to connect to
    """
    for line in file:
        item = create_item(line)
        item.experiment_id = experiment
        item.save()
        for _id, level in enumerate(line.levels):
            itemlevel = ItemLevel()
            itemlevel.item = item
            for schema_level in levels:
                if schema_level[1] == level:
                    itemlevel.level = schema_level[0]
                    itemlevel.save()
                    break


def insert_factors(schema: dict, experiment_id: Experiment) -> [(Level, str, str)]:
    """
    inserts the factors and levels to the database
    :param schema: the input structure of levels and factors from the web
    :param experiment_id: the created Experiment
    :return: list with level reference, name and factor name
    """
    levels = []
    for factor in schema:
        db_factor = Factor(name=factor["name"], experiment_id=experiment_id)
        db_factor.save()
        for level in factor["level"]:
            db_level = Level(name=level["name"], factor=db_factor)
            db_level.save()
            levels.append((db_level, db_level.name, db_factor.name))
    return levels
